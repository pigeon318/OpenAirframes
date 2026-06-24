import os
import sys
from datetime import datetime
from pathlib import Path

import psycopg
from dotenv import load_dotenv
from openpyxl import load_workbook

script_dir = Path(__file__).parent
data_dir = script_dir.parent / "data"
SOURCE = "UK_CAA"
TRUST_ORDER = ["FAA", "UK_CAA", "Mictronics", "OpenSky", "user"]

COLUMN_MAP = {
    "registration": ["registration marks", "mark", "registration"],
    "icao_hex": ["icao 24 bit address", "mode s code (hex)", "icao"],
    "manufacturer": ["manufacturer", "aircraft type/manufacturer", "manufacturer name"],
    "model": ["generic name", "popular name", "designation", "model"],
    "serial_number": ["serial no", "serial number", "aircraft serial no"],
    "owner_name": ["registered owner", "owner", "full names"],
    "ownership_status": ["ownership status"],
    "aircraft_class": ["aircraft class", "class"],
    "engine_count": ["number of engines", "no of engines", "engines"],
    "engine_type": ["engine type", "engine"],
    "engine_manufacturer": ["engine manufacturer"],
    "engine_class": ["engine class"],
    "cert_category": ["airworthiness certificate category", "coa category", "certificate category"],
    "cert_expiry": ["airworthiness certificate expiry", "coa expiry", "certificate expiry"],
    "mtow": ["maximum take off weight", "mtow", "max take off weight"],
    "year_of_construction": ["year of construction", "year built", "year manufactured"],
    "max_passengers": ["maximum number of passengers", "max passengers", "seats"],
    "date_current_reg": ["date of current registration", "current registration date"],
    "date_first_reg": ["date of first registration", "first registration date"],
    "previous_identity": ["previous identity", "previous registration"],
    "address_1": ["address 1", "registered address 1"],
    "address_2": ["address 2", "registered address 2"],
    "address_3": ["address 3", "registered address 3"],
    "address_4": ["address 4", "registered address 4"],
    "address_5": ["address 5", "registered address 5"],
}


def normalise_header(h):
    return str(h).strip().lower() if h is not None else ""


def build_header_index(headers):
    norm = {normalise_header(h): i for i, h in enumerate(headers)}
    index = {}
    for field, candidates in COLUMN_MAP.items():
        for candidate in candidates:
            if candidate in norm:
                index[field] = norm[candidate]
                break
    return index


def cell_str(cell):
    if cell is None or cell.value is None:
        return None
    v = str(cell.value).strip()
    return v if v else None


def cell_date(cell):
    if cell is None or cell.value is None:
        return None
    if isinstance(cell.value, datetime):
        return cell.value.date()
    s = str(cell.value).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def cell_int(cell):
    if cell is None or cell.value is None:
        return None
    try:
        return int(cell.value)
    except (ValueError, TypeError):
        try:
            return int(str(cell.value).strip())
        except (ValueError, TypeError):
            return None


def get_col(row, index, field):
    i = index.get(field)
    if i is None:
        return None
    return row[i] if i < len(row) else None


def load_register(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows()
    header_row = next(rows, None)
    if header_row is None:
        return

    headers = [c.value for c in header_row]
    index = build_header_index(headers)

    for row in rows:
        icao_cell = get_col(row, index, "icao_hex")
        icao_raw = cell_str(icao_cell)
        if not icao_raw:
            continue
        icao_hex = icao_raw.lower()
        if len(icao_hex) != 6 or not all(c in "0123456789abcdef" for c in icao_hex):
            continue

        reg_cell = get_col(row, index, "registration")
        registration = cell_str(reg_cell)

        addr_parts = [
            cell_str(get_col(row, index, f"address_{i}"))
            for i in range(1, 6)
        ]
        addr_parts = [p for p in addr_parts if p]
        owner_city = addr_parts[-3] if len(addr_parts) >= 3 else (addr_parts[0] if addr_parts else None)
        owner_country = addr_parts[-1] if addr_parts else None

        engine_parts = [
            cell_str(get_col(row, index, "engine_manufacturer")),
            cell_str(get_col(row, index, "engine_type")),
            cell_str(get_col(row, index, "engine_class")),
        ]
        engine_type_raw = " ".join(p for p in engine_parts if p) or None

        cert_raw = cell_str(get_col(row, index, "cert_category"))

        yield {
            "icao_hex": icao_hex,
            "registration": registration,
            "serial_number": cell_str(get_col(row, index, "serial_number")),
            "source_record_id": None,
            "manufacturer": cell_str(get_col(row, index, "manufacturer")),
            "model": cell_str(get_col(row, index, "model")),
            "type_aircraft": None,
            "type_aircraft_raw": cell_str(get_col(row, index, "aircraft_class")),
            "type_engine": None,
            "type_engine_raw": engine_type_raw,
            "engine_count": cell_int(get_col(row, index, "engine_count")),
            "seats": cell_int(get_col(row, index, "max_passengers")),
            "year_manufactured": cell_int(get_col(row, index, "year_of_construction")),
            "owner_name": cell_str(get_col(row, index, "owner_name")),
            "owner_type": cell_str(get_col(row, index, "ownership_status")),
            "owner_type_raw": cell_str(get_col(row, index, "ownership_status")),
            "owner_city": owner_city,
            "owner_state": None,
            "owner_country": owner_country,
            "status": "Valid",
            "status_raw": None,
            "certification": cert_raw,
            "last_action_date": cell_date(get_col(row, index, "date_current_reg")),
            "cert_issue_date": cell_date(get_col(row, index, "date_first_reg")),
            "airworthiness_date": cell_date(get_col(row, index, "date_first_reg")),
            "expiration_date": cell_date(get_col(row, index, "cert_expiry")),
            "aircraft_category": None,
            "aircraft_category_raw": None,
            "builder_certification": None,
            "builder_certification_raw": None,
            "weight_class": None,
            "weight_class_raw": cell_str(get_col(row, index, "mtow")),
            "source": SOURCE,
        }


COLUMNS = [
    "icao_hex",
    "registration",
    "serial_number",
    "source_record_id",
    "manufacturer",
    "model",
    "type_aircraft",
    "type_aircraft_raw",
    "type_engine",
    "type_engine_raw",
    "engine_count",
    "seats",
    "year_manufactured",
    "owner_name",
    "owner_type",
    "owner_type_raw",
    "owner_city",
    "owner_state",
    "owner_country",
    "status",
    "status_raw",
    "certification",
    "last_action_date",
    "cert_issue_date",
    "airworthiness_date",
    "expiration_date",
    "aircraft_category",
    "aircraft_category_raw",
    "builder_certification",
    "builder_certification_raw",
    "weight_class",
    "weight_class_raw",
    "source",
]


def ingest(conn, path):
    inserted = 0
    skipped_dup = 0
    seen = set()

    trust_case = " ".join(
        f"WHEN '{src}' THEN {len(TRUST_ORDER) - i}"
        for i, src in enumerate(TRUST_ORDER)
    )

    with conn.cursor() as cur:
        cur.execute("""
            CREATE TEMP TABLE aircraft_staging (
                LIKE aircraft INCLUDING DEFAULTS
            ) ON COMMIT DROP
        """)

        copy_sql = f"COPY aircraft_staging ({', '.join(COLUMNS)}) FROM STDIN"
        with cur.copy(copy_sql) as copy:
            for record in load_register(path):
                if record["icao_hex"] in seen:
                    skipped_dup += 1
                    continue
                seen.add(record["icao_hex"])
                copy.write_row([record[c] for c in COLUMNS])
                inserted += 1

        upsert_sql = f"""
            INSERT INTO aircraft ({', '.join(COLUMNS)})
            SELECT {', '.join(COLUMNS)} FROM aircraft_staging
            ON CONFLICT (icao_hex) DO UPDATE SET
                registration = EXCLUDED.registration,
                serial_number = EXCLUDED.serial_number,
                source_record_id = EXCLUDED.source_record_id,
                manufacturer = EXCLUDED.manufacturer,
                model = EXCLUDED.model,
                type_aircraft = COALESCE(EXCLUDED.type_aircraft, aircraft.type_aircraft),
                type_aircraft_raw = EXCLUDED.type_aircraft_raw,
                type_engine = COALESCE(EXCLUDED.type_engine, aircraft.type_engine),
                type_engine_raw = EXCLUDED.type_engine_raw,
                engine_count = COALESCE(EXCLUDED.engine_count, aircraft.engine_count),
                seats = COALESCE(EXCLUDED.seats, aircraft.seats),
                year_manufactured = COALESCE(EXCLUDED.year_manufactured, aircraft.year_manufactured),
                owner_name = EXCLUDED.owner_name,
                owner_type = EXCLUDED.owner_type,
                owner_type_raw = EXCLUDED.owner_type_raw,
                owner_city = EXCLUDED.owner_city,
                owner_state = EXCLUDED.owner_state,
                owner_country = EXCLUDED.owner_country,
                status = EXCLUDED.status,
                status_raw = EXCLUDED.status_raw,
                certification = EXCLUDED.certification,
                last_action_date = EXCLUDED.last_action_date,
                cert_issue_date = EXCLUDED.cert_issue_date,
                airworthiness_date = EXCLUDED.airworthiness_date,
                expiration_date = EXCLUDED.expiration_date,
                aircraft_category = COALESCE(EXCLUDED.aircraft_category, aircraft.aircraft_category),
                aircraft_category_raw = EXCLUDED.aircraft_category_raw,
                builder_certification = COALESCE(EXCLUDED.builder_certification, aircraft.builder_certification),
                builder_certification_raw = EXCLUDED.builder_certification_raw,
                weight_class = COALESCE(EXCLUDED.weight_class, aircraft.weight_class),
                weight_class_raw = EXCLUDED.weight_class_raw,
                source = EXCLUDED.source,
                updated_at = NOW()
            WHERE (CASE aircraft.source {trust_case} END)
                  <= (CASE EXCLUDED.source {trust_case} END)
        """
        cur.execute(upsert_sql)
        affected = cur.rowcount

    conn.commit()
    return inserted, skipped_dup, affected


def print_columns(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(), None)
    if not header_row:
        print("No rows found in file.")
        return
    print("Columns found in file:")
    for i, cell in enumerate(header_row):
        print(f"  [{i}] {cell.value!r}")
    index = build_header_index([c.value for c in header_row])
    print("\nMapped to schema fields:")
    for field, col_i in index.items():
        print(f"  {field} -> column {col_i} ({header_row[col_i].value!r})")
    unmapped = [f for f in COLUMN_MAP if f not in index]
    if unmapped:
        print("\nUnmapped fields (will be null):")
        for f in unmapped:
            print(f"  {f}")


def main():
    load_dotenv(script_dir.parent / ".env")

    xl_path = data_dir / "GINFO.xlsx"
    flag = None

    args = sys.argv[1:]
    if args and args[0] == "--print-columns":
        flag = "print-columns"
        args = args[1:]
    if args:
        xl_path = Path(args[0])

    if not xl_path.exists():
        print(f"Register file not found: {xl_path}", file=sys.stderr)
        print("Expected path: data/GINFO.xlsx", file=sys.stderr)
        sys.exit(1)

    if flag == "print-columns":
        print_columns(xl_path)
        return

    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        print("DATABASE_URL not set", file=sys.stderr)
        sys.exit(1)

    started = datetime.now()
    with psycopg.connect(database_url) as conn:
        inserted, skipped, affected = ingest(conn, xl_path)

    elapsed = (datetime.now() - started).total_seconds()
    print(f"UK CAA ingest complete in {elapsed:.1f}s")
    print(f"  staged: {inserted}")
    if skipped:
        print(f"  duplicate icao_hex skipped: {skipped}")
    print(f"  inserted or updated: {affected}")


if __name__ == "__main__":
    main()
